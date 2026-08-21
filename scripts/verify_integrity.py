import os
import pandas as pd
import openpyxl

def verify_project_integrity():
    print("=== STARTING INTEGRITY VERIFICATION ===")
    
    # 1. Check file existence
    paths = {
        "Raw Excel Data": os.path.join("data", "raw_sales_data.xlsx"),
        "Cleaned Excel Data": os.path.join("data", "cleaned_sales_data.xlsx"),
        "Excel Analysis Workbook": os.path.join("excel", "Sales_Analysis.xlsx"),
        "Power BI File": os.path.join("powerbi", "Sales_Performance_Dashboard.pbix"),
        "Dashboard Screenshot": os.path.join("screenshots", "dashboard.png"),
        "Business Insights Doc": os.path.join("documentation", "business_insights.md"),
        "Data Dictionary Doc": os.path.join("documentation", "data_dictionary.md"),
        "Data Cleaning Doc": os.path.join("documentation", "data_cleaning.md"),
        "DAX Measures Doc": os.path.join("documentation", "dax_measures.md"),
    }
    
    missing_files = 0
    for name, path in paths.items():
        if os.path.exists(path):
            print(f"[PASSED] File exists: {name} ({path})")
        else:
            print(f"[FAILED] Missing file: {name} ({path})")
            missing_files += 1
            
    # 2. Check Data Row Counts
    df_raw = pd.read_excel(paths["Raw Excel Data"])
    df_clean = pd.read_excel(paths["Cleaned Excel Data"])
    
    print(f"Raw data rows: {len(df_raw)}")
    print(f"Cleaned data rows: {len(df_clean)}")
    
    # 3. Check Excel workbook tabs and row counts
    try:
        wb = openpyxl.load_workbook(paths["Excel Analysis Workbook"], read_only=True)
        sheets = wb.sheetnames
        print(f"Excel workbook sheets found: {sheets}")
        
        expected_sheets = ["Raw Data", "Cleaned Data", "Pivot Tables", "Dashboard"]
        for s in expected_sheets:
            if s in sheets:
                print(f"[PASSED] Worksheet found: '{s}'")
            else:
                print(f"[FAILED] Missing worksheet: '{s}'")
                
        # Check rows in Cleaned Data sheet of final workbook
        ws_clean = wb["Cleaned Data"]
        row_count = ws_clean.max_row
        print(f"Cleaned Data sheet row count (with header): {row_count}")
        # max_row in read-only openpyxl returns the correct last row
        if row_count - 1 == len(df_clean):
            print(f"[PASSED] Cleaned data row counts match exactly: {len(df_clean)} rows.")
        else:
            print(f"[WARNING] Row count mismatch: excel sheet has {row_count - 1} rows but cleaned data file has {len(df_clean)} rows.")
            
        wb.close()
    except Exception as e:
        print(f"[FAILED] Error reading Excel workbook: {e}")
        
    # 4. Math verification
    pandas_sales = df_clean["Sales"].sum()
    pandas_profit = df_clean["Profit"].sum()
    
    print(f"Total Sales (Pandas): ${pandas_sales:,.2f}")
    print(f"Total Profit (Pandas): ${pandas_profit:,.2f}")
    
    # Verify that totals match the ones written to our insights document
    with open(paths["Business Insights Doc"], "r") as f:
        insights_content = f.read()
        
    if f"${pandas_sales:,.2f}" in insights_content:
        print("[PASSED] Total Sales in insights doc matches calculated sales.")
    else:
        print("[FAILED] Total Sales mismatch in insights doc.")
        
    if f"${pandas_profit:,.2f}" in insights_content:
        print("[PASSED] Total Profit in insights doc matches calculated profit.")
    else:
        print("[FAILED] Total Profit mismatch in insights doc.")
        
    print("=== INTEGRITY VERIFICATION COMPLETE ===")
    
if __name__ == "__main__":
    verify_project_integrity()
